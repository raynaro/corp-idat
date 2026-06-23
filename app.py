import os
import sqlite3
from zoneinfo import ZoneInfo
os.makedirs(UPLOADS, exist_ok=True)
from datetime import datetime, timedelta, time
from functools import wraps
from zoneinfo import ZoneInfo

from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

APP_DIR = os.path.dirname(os.path.abspath(__file__))

DB_PATH = os.path.join(APP_DIR, "database.db")

UPLOADS = os.path.join(APP_DIR, "uploads")
BASE_XLSX = os.path.join(APP_DIR, "base.xlsx")
TZ = ZoneInfo("America/Lima")

os.makedirs(UPLOADS, exist_ok=True)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'cambia-esta-clave-en-produccion')

ROLES = {
    'admin': 'Administrador',
    'sae': 'SAE',
    'sales': 'Ventas',
    'security': 'Seguridad'
}
AREAS = ('SAE', 'VENTAS')


def conn():
    c = sqlite3.connect(DB_NAME)
    c.row_factory = sqlite3.Row
    return c


def now_dt():
    return datetime.now(TZ)


def now_iso():
    return now_dt().isoformat(timespec='seconds')


def today_lima():
    return now_dt().date().isoformat()


def purge_old_logs():
    # Conserva registros operativos 7 días para no crecer demasiado.
    # Las visitas nuevas también se conservan 7 días porque el reporte es semanal.
    limite = (now_dt() - timedelta(days=30)).isoformat(timespec='seconds')
    c = conn()
    c.execute('DELETE FROM searches WHERE created_at < ?', (limite,))
    c.execute('DELETE FROM access_logs WHERE created_at < ?', (limite,))
    c.execute('DELETE FROM teacher_logs WHERE created_at < ?', (limite,))
    c.execute('DELETE FROM visitors WHERE created_at < ?', (limite,))
    c.commit(); c.close()


def normalize_header(v):
    return str(v or '').strip().lower().replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')


def only_digits(v):
    return str(v or '').strip().replace('.0', '')


def import_students_from_excel(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [normalize_header(c.value) for c in ws[1]]
    def find(*names):
        names = [normalize_header(n) for n in names]
        for n in names:
            if n in headers:
                return headers.index(n)
        return None
    i_name = find('NombreCompleto', 'Nombre Completo', 'Nombre', 'Alumno')
    i_dni = find('DNI', 'Documento')
    i_code = find('Codigo', 'Código', 'Code')
    i_date = find('Fecha Ingreso', 'Fecha de Ingreso', 'Fecha')
    if i_name is None or i_dni is None or i_code is None:
        raise ValueError('El Excel debe tener columnas: NombreCompleto, DNI y Codigo')

    c = conn(); cur = c.cursor(); count = 0
    cur.execute('UPDATE students SET active=0, updated_at=?', (now_iso(),))
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[i_name] or '').strip().upper()
        dni = only_digits(row[i_dni])
        code = only_digits(row[i_code]).upper()
        entry_date = str(row[i_date] or '').strip() if i_date is not None else now_dt().strftime('%d/%m/%Y')
        if not name or not dni or not code:
            continue
        cur.execute('''INSERT INTO students(name,dni,code,entry_date,active,created_at,updated_at)
                       VALUES(?,?,?,?,1,?,?)
                       ON CONFLICT(code) DO UPDATE SET
                       name=excluded.name, dni=excluded.dni, entry_date=excluded.entry_date,
                       active=1, updated_at=excluded.updated_at''',
                    (name, dni, code, entry_date, now_iso(), now_iso()))
        count += 1
    c.commit(); c.close()
    return count


def import_teachers_from_excel(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [normalize_header(c.value) for c in ws[1]]
    def find(*names):
        names = [normalize_header(n) for n in names]
        for n in names:
            if n in headers:
                return headers.index(n)
        return None
    i_name = find('NombreCompleto', 'Nombre Completo', 'Nombre', 'Docente')
    i_dni = find('DNI', 'Documento')
    i_area = find('Area', 'Área', 'Especialidad', 'Curso')
    if i_name is None or i_dni is None:
        raise ValueError('El Excel de docentes debe tener columnas: NombreCompleto y DNI. Área/Especialidad es opcional.')

    c = conn(); cur = c.cursor(); count = 0
    cur.execute('UPDATE teachers SET active=0, updated_at=?', (now_iso(),))
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[i_name] or '').strip().upper()
        dni = only_digits(row[i_dni])
        area = str(row[i_area] or '').strip().upper() if i_area is not None else ''
        if not name or not dni:
            continue
        cur.execute('''INSERT INTO teachers(name,dni,area,active,created_at,updated_at)
                       VALUES(?,?,?,1,?,?)
                       ON CONFLICT(dni) DO UPDATE SET
                       name=excluded.name, area=excluded.area, active=1, updated_at=excluded.updated_at''',
                    (name, dni, area, now_iso(), now_iso()))
        count += 1
    c.commit(); c.close()
    return count


def ensure_column(cur, table, column, definition):
    cols = [r['name'] for r in cur.execute(f'PRAGMA table_info({table})').fetchall()]
    if column not in cols:
        cur.execute(f'ALTER TABLE {table} ADD COLUMN {column} {definition}')


def init_db():
    c = conn(); cur = c.cursor()
    cur.execute('''CREATE TABLE IF NOT EXISTS users(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        email TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'security',
        created_at TEXT NOT NULL
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS students(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        dni TEXT NOT NULL,
        code TEXT NOT NULL UNIQUE,
        entry_date TEXT,
        active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS teachers(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        dni TEXT NOT NULL UNIQUE,
        area TEXT,
        active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS searches(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        query TEXT NOT NULL,
        result TEXT NOT NULL,
        student_name TEXT,
        created_at TEXT NOT NULL
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS access_logs(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        student_id INTEGER,
        query TEXT NOT NULL,
        result TEXT NOT NULL,
        student_name TEXT,
        student_dni TEXT,
        student_code TEXT,
        note TEXT,
        created_at TEXT NOT NULL
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS teacher_logs(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        teacher_id INTEGER,
        query TEXT NOT NULL,
        result TEXT NOT NULL,
        teacher_name TEXT,
        teacher_dni TEXT,
        teacher_area TEXT,
        note TEXT,
        created_at TEXT NOT NULL
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS visitors(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        full_name TEXT NOT NULL,
        dni TEXT NOT NULL,
        destination_area TEXT NOT NULL,
        reason TEXT,
        attended INTEGER NOT NULL DEFAULT 0,
        attended_at TEXT,
        created_at TEXT NOT NULL
    )''')
    ensure_column(cur, 'users', 'role', "TEXT NOT NULL DEFAULT 'security'")
    ensure_column(cur, 'visitors', 'attended', "INTEGER NOT NULL DEFAULT 0")
    ensure_column(cur, 'visitors', 'attended_at', "TEXT")

    users = [
        ('Administrador', 'admin@idat.edu.pe', 'admin123', 'admin'),
        ('Personal SAE', 'sae@idat.edu.pe', 'sae123', 'sae'),
        ('Personal Ventas', 'ventas@idat.edu.pe', 'ventas123', 'sales'),
        ('Personal de Seguridad', 'seguridad@idat.edu.pe', 'seguridad123', 'security')
    ]
    for name,email,pwd,role in users:
        u = cur.execute('SELECT id FROM users WHERE email=?', (email,)).fetchone()
        if not u:
            cur.execute('INSERT INTO users(name,email,password,role,created_at) VALUES(?,?,?,?,?)',
                        (name,email,generate_password_hash(pwd),role,now_iso()))
    c.commit(); c.close()
    c = conn(); total = c.execute('SELECT COUNT(*) n FROM students').fetchone()['n']; c.close()
    if total == 0 and os.path.exists(BASE_XLSX):
        import_students_from_excel(BASE_XLSX)
    purge_old_logs()


def login_required(f):
    @wraps(f)
    def w(*a, **kw):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        purge_old_logs()
        return f(*a, **kw)
    return w


def role_required(*roles):
    def deco(f):
        @wraps(f)
        def w(*a, **kw):
            if session.get('role') not in roles:
                flash('No tienes permiso para entrar a esa sección.', 'danger')
                return redirect(url_for('consulta'))
            return f(*a, **kw)
        return w
    return deco


def admin_required(f):
    return role_required('admin')(f)


@app.context_processor
def inject_globals():
    return dict(roles=ROLES, areas=AREAS)


@app.route('/')
def home():
    return redirect(url_for('login'))


@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email','').strip().lower()
        password = request.form.get('password','')
        c = conn(); u = c.execute('SELECT * FROM users WHERE email=?', (email,)).fetchone(); c.close()
        if u and check_password_hash(u['password'], password):
            session.clear(); session['user_id']=u['id']; session['name']=u['name']; session['role']=u['role']
            
            if u['role'] == 'admin':
                return redirect(url_for('dashboard'))
            if u['role'] in ('sae','sales'):
                return redirect(url_for('visitors'))
            if u['role'] == 'security':
                return redirect(url_for('security_panel'))
            return redirect(url_for('consulta'))
        flash('Credenciales incorrectas.', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear(); return redirect(url_for('login'))


@app.route('/register', methods=['GET','POST'])
@login_required
@admin_required
def register():
    if request.method == 'POST':
        name=request.form.get('name','').strip(); email=request.form.get('email','').strip().lower(); password=request.form.get('password','')
        role = request.form.get('role','security')
        if role not in ROLES: role = 'security'
        if not name or not email or not password:
            flash('Completa todos los campos.', 'danger'); return render_template('register.html')
        try:
            c=conn(); c.execute('INSERT INTO users(name,email,password,role,created_at) VALUES(?,?,?,?,?)',
                (name,email,generate_password_hash(password),role,now_iso())); c.commit(); c.close()
            flash('Usuario creado correctamente.', 'success')
            return redirect(url_for('users'))
        except sqlite3.IntegrityError:
            flash('Ese correo ya existe.', 'danger')
    return render_template('register.html')



@app.route('/seguridad')
@login_required
@role_required('admin','security')
def security_panel():
    c = conn(); today = today_lima()
    stats = {
        'students_today': c.execute("SELECT COUNT(*) n FROM access_logs WHERE user_id=? AND substr(created_at,1,10)=?", (session['user_id'], today)).fetchone()['n'],
        'teachers_today': c.execute("SELECT COUNT(*) n FROM teacher_logs WHERE user_id=? AND substr(created_at,1,10)=?", (session['user_id'], today)).fetchone()['n'],
        'visits_today': c.execute("SELECT COUNT(*) n FROM visitors WHERE user_id=? AND substr(created_at,1,10)=?", (session['user_id'], today)).fetchone()['n'],
        'lima_now': now_dt().strftime('%d/%m/%Y %H:%M')
    }
    recent_students = c.execute('''SELECT * FROM access_logs WHERE user_id=? ORDER BY id DESC LIMIT 5''', (session['user_id'],)).fetchall()
    recent_teachers = c.execute('''SELECT * FROM teacher_logs WHERE user_id=? ORDER BY id DESC LIMIT 5''', (session['user_id'],)).fetchall()
    recent_visits = c.execute('''SELECT * FROM visitors WHERE user_id=? ORDER BY id DESC LIMIT 5''', (session['user_id'],)).fetchall()
    c.close()
    return render_template('security_panel.html', stats=stats, recent_students=recent_students, recent_teachers=recent_teachers, recent_visits=recent_visits)

@app.route('/consulta', methods=['GET','POST'])
@login_required
def consulta():
    result = None; q=''; status=None
    if request.method == 'POST':
        q = request.form.get('q','').strip().upper()
        note = request.form.get('note','').strip()
        c=conn()
        st = c.execute('SELECT * FROM students WHERE dni=? OR code=?', (q,q)).fetchone()
        status = 'ACTIVO' if st and st['active'] else 'INACTIVO'
        student_name = st['name'] if st else None
        created = now_iso()
        c.execute('INSERT INTO searches(user_id,query,result,student_name,created_at) VALUES(?,?,?,?,?)',
                  (session['user_id'],q,status,student_name,created))
        c.execute('''INSERT INTO access_logs(user_id,student_id,query,result,student_name,student_dni,student_code,note,created_at)
                     VALUES(?,?,?,?,?,?,?,?,?)''',
                  (session['user_id'], st['id'] if st else None, q, status, student_name,
                   st['dni'] if st else None, st['code'] if st else None, note, created))
        c.commit(); c.close()
        result = st
    return render_template('consulta.html', result=result, q=q, status=status)


@app.route('/visitas/nueva', methods=['GET','POST'])
@login_required
@role_required('admin','sae','sales','security')
def new_visitor():
    if request.method == 'POST':
        full_name = request.form.get('full_name','').strip().upper()
        dni = only_digits(request.form.get('dni',''))
        destination_area = request.form.get('destination_area','').strip().upper()
        if session.get('role') == 'sae':
            destination_area = 'SAE'
        elif session.get('role') == 'sales':
            destination_area = 'VENTAS'
        reason = request.form.get('reason','').strip()
        if destination_area not in AREAS:
            flash('Selecciona el área SAE o VENTAS.', 'danger'); return render_template('visitor_form.html')
        if not full_name or not dni:
            flash('Completa nombres completos y DNI.', 'danger'); return render_template('visitor_form.html')
        c=conn()
        c.execute('INSERT INTO visitors(user_id,full_name,dni,destination_area,reason,created_at) VALUES(?,?,?,?,?,?)',
                  (session['user_id'], full_name, dni, destination_area, reason, now_iso()))
        c.commit(); c.close()
        flash('Visita registrada correctamente.', 'success')
        if session.get('role') == 'security':
            return redirect(url_for('security_panel'))
        return redirect(url_for('visitors'))
    selected_area = request.args.get('area','').strip().upper()
    if selected_area not in AREAS:
        selected_area = ''
    return render_template('visitor_form.html', selected_area=selected_area)



@app.route('/sae')
@login_required
@role_required('admin','sae')
def sae_area():
    session['area_view'] = 'SAE'
    return redirect(url_for('visitors'))


@app.route('/ventas')
@login_required
@role_required('admin','sales')
def sales_area():
    session['area_view'] = 'VENTAS'
    return redirect(url_for('visitors'))



@app.route('/docentes/registro', methods=['GET','POST'])
@login_required
@role_required('admin','security')
def teacher_check():
    result = None; q=''; status=None
    if request.method == 'POST':
        q = request.form.get('q','').strip().upper()
        note = request.form.get('note','').strip()
        c=conn()
        teacher = c.execute('SELECT * FROM teachers WHERE dni=?', (q,)).fetchone()
        status = 'ACTIVO' if teacher and teacher['active'] else 'INACTIVO'
        created = now_iso()
        c.execute('''INSERT INTO teacher_logs(user_id,teacher_id,query,result,teacher_name,teacher_dni,teacher_area,note,created_at)
                     VALUES(?,?,?,?,?,?,?,?,?)''',
                  (session['user_id'], teacher['id'] if teacher else None, q, status,
                   teacher['name'] if teacher else None, teacher['dni'] if teacher else None,
                   teacher['area'] if teacher else None, note, created))
        c.commit(); c.close()
        result = teacher
    return render_template('teacher_check.html', result=result, q=q, status=status)

@app.route('/visitas')
@login_required
@role_required('admin','sae','sales','security')
def visitors():
    c=conn()
    role = session.get('role')
    area_actual = None
    sonido_area = False
    if role == 'sae':
        area_actual = 'SAE'
        sonido_area = True
        data = c.execute('''SELECT v.*, u.name user FROM visitors v LEFT JOIN users u ON u.id=v.user_id
                            WHERE v.destination_area='SAE' ORDER BY v.id ASC''').fetchall()
    elif role == 'sales':
        area_actual = 'VENTAS'
        sonido_area = True
        data = c.execute('''SELECT v.*, u.name user FROM visitors v LEFT JOIN users u ON u.id=v.user_id
                            WHERE v.destination_area='VENTAS' ORDER BY v.id ASC''').fetchall()
    elif role == 'security':
        data = c.execute('''SELECT v.*, u.name user FROM visitors v LEFT JOIN users u ON u.id=v.user_id
                            WHERE v.user_id=? ORDER BY v.id DESC''', (session['user_id'],)).fetchall()
    else:
        data = c.execute('''SELECT v.*, u.name user FROM visitors v LEFT JOIN users u ON u.id=v.user_id ORDER BY v.id DESC''').fetchall()
    ultimo = data[-1]['id'] if data and role in ('sae','sales') else 0
    c.close()
    return render_template('visitors.html', visitors=data, area_actual=area_actual, sonido_area=sonido_area, ultimo_id=ultimo)


@app.route('/api/visitas_estado')
@login_required
@role_required('admin','sae','sales')
def visitas_estado():
    role = session.get('role')
    area = None
    if role == 'sae':
        area = 'SAE'
    elif role == 'sales':
        area = 'VENTAS'
    else:
        area = request.args.get('area','').strip().upper()
        if area not in AREAS:
            area = None
    try:
        since_id = int(request.args.get('since', '0') or 0)
    except ValueError:
        since_id = 0
    c = conn()
    if area:
        row = c.execute("""SELECT COUNT(*) total,
                                  COALESCE(MAX(id),0) ultimo,
                                  SUM(CASE WHEN attended=0 THEN 1 ELSE 0 END) pendientes,
                                  COALESCE(MAX(CASE WHEN attended=0 THEN id ELSE 0 END),0) ultimo_pendiente
                           FROM visitors WHERE destination_area=?""", (area,)).fetchone()
        nuevos = c.execute('SELECT COUNT(*) n FROM visitors WHERE destination_area=? AND id>?', (area, since_id)).fetchone()['n']
    else:
        row = c.execute("""SELECT COUNT(*) total,
                                  COALESCE(MAX(id),0) ultimo,
                                  SUM(CASE WHEN attended=0 THEN 1 ELSE 0 END) pendientes,
                                  COALESCE(MAX(CASE WHEN attended=0 THEN id ELSE 0 END),0) ultimo_pendiente
                           FROM visitors""").fetchone()
        nuevos = c.execute('SELECT COUNT(*) n FROM visitors WHERE id>?', (since_id,)).fetchone()['n']
    c.close()
    return jsonify({
        'total': row['total'] or 0,
        'ultimo': row['ultimo'] or 0,
        'pendientes': row['pendientes'] or 0,
        'ultimo_pendiente': row['ultimo_pendiente'] or 0,
        'nuevos': nuevos or 0,
        'area': area or 'TODAS'
    })

@app.route('/visitas/<int:visit_id>/estado', methods=['POST'])
@login_required
@role_required('admin','sae','sales')
def update_visit_status(visit_id):
    role = session.get('role')
    estado = request.form.get('estado', '0')
    attended = 1 if estado == '1' else 0
    c = conn()
    visit = c.execute('SELECT * FROM visitors WHERE id=?', (visit_id,)).fetchone()
    if not visit:
        c.close(); flash('La visita no existe.', 'danger'); return redirect(url_for('visitors'))
    if role == 'sae' and visit['destination_area'] != 'SAE':
        c.close(); flash('No tienes permiso para cambiar esta visita.', 'danger'); return redirect(url_for('visitors'))
    if role == 'sales' and visit['destination_area'] != 'VENTAS':
        c.close(); flash('No tienes permiso para cambiar esta visita.', 'danger'); return redirect(url_for('visitors'))
    c.execute('UPDATE visitors SET attended=?, attended_at=? WHERE id=?',
              (attended, now_iso() if attended else None, visit_id))
    c.commit(); c.close()
    flash('Estado de atención actualizado.', 'success')
    return redirect(url_for('visitors'))




def week_dates_lima(days=7):
    base = now_dt().date()
    return [(base - timedelta(days=i)).isoformat() for i in range(days-1, -1, -1)]

@app.route('/dashboard')
@login_required
@admin_required
def dashboard():
    c=conn(); today=today_lima()
    week_days = week_dates_lima(7)
    stats = {
        'active': c.execute('SELECT COUNT(*) n FROM students WHERE active=1').fetchone()['n'],
        'teachers': c.execute('SELECT COUNT(*) n FROM teachers WHERE active=1').fetchone()['n'],
        'today': c.execute("SELECT COUNT(*) n FROM access_logs WHERE substr(created_at,1,10)=?", (today,)).fetchone()['n'],
        'visitors': c.execute("SELECT COUNT(*) n FROM visitors WHERE substr(created_at,1,10)=?", (today,)).fetchone()['n'],
        'users': c.execute('SELECT COUNT(*) n FROM users').fetchone()['n'],
        'inactive': c.execute('SELECT COUNT(*) n FROM access_logs WHERE result="INACTIVO"').fetchone()['n'],
        'visitors_week': c.execute("SELECT COUNT(*) n FROM visitors WHERE substr(created_at,1,10) BETWEEN ? AND ?", (week_days[0], week_days[-1])).fetchone()['n'],
        'pending': c.execute('SELECT COUNT(*) n FROM visitors WHERE attended=0').fetchone()['n'],
        'attended': c.execute('SELECT COUNT(*) n FROM visitors WHERE attended=1').fetchone()['n'],
        'students_active_today': c.execute("SELECT COUNT(*) n FROM access_logs WHERE result='ACTIVO' AND substr(created_at,1,10)=?", (today,)).fetchone()['n'],
        'teachers_active_today': c.execute("SELECT COUNT(*) n FROM teacher_logs WHERE result='ACTIVO' AND substr(created_at,1,10)=?", (today,)).fetchone()['n']
    }
    weekly_rows = c.execute("""SELECT substr(created_at,1,10) dia, COUNT(*) total,
                                      SUM(CASE WHEN destination_area='SAE' THEN 1 ELSE 0 END) sae,
                                      SUM(CASE WHEN destination_area='VENTAS' THEN 1 ELSE 0 END) ventas
                               FROM visitors
                               WHERE substr(created_at,1,10) BETWEEN ? AND ?
                               GROUP BY substr(created_at,1,10)""", (week_days[0], week_days[-1])).fetchall()
    weekly_map = {r['dia']: r for r in weekly_rows}
    weekly = []
    max_total = 1
    for d in week_days:
        r = weekly_map.get(d)
        item = {'dia': d[5:], 'total': (r['total'] if r else 0), 'sae': (r['sae'] if r else 0), 'ventas': (r['ventas'] if r else 0)}
        max_total = max(max_total, item['total'])
        weekly.append(item)
    for item in weekly:
        item['pct'] = int((item['total'] / max_total) * 100) if max_total else 0

    area_stats = c.execute("""SELECT destination_area area, COUNT(*) total,
                                     SUM(CASE WHEN attended=1 THEN 1 ELSE 0 END) atendidos,
                                     SUM(CASE WHEN attended=0 THEN 1 ELSE 0 END) pendientes
                              FROM visitors GROUP BY destination_area ORDER BY destination_area""").fetchall()
    sae_detail = c.execute("""SELECT v.*, u.name user FROM visitors v LEFT JOIN users u ON u.id=v.user_id
                              WHERE v.destination_area='SAE' ORDER BY v.id DESC LIMIT 12""").fetchall()
    sales_detail = c.execute("""SELECT v.*, u.name user FROM visitors v LEFT JOIN users u ON u.id=v.user_id
                                WHERE v.destination_area='VENTAS' ORDER BY v.id DESC LIMIT 12""").fetchall()
    security_detail = c.execute("""SELECT v.*, u.name user FROM visitors v LEFT JOIN users u ON u.id=v.user_id
                                   WHERE u.role='security' ORDER BY v.id DESC LIMIT 12""").fetchall()
    user_activity = c.execute("""
        SELECT u.name, u.role,
               COALESCE(a.alumnos,0) alumnos,
               COALESCE(t.docentes,0) docentes,
               COALESCE(v.visitas,0) visitas,
               COALESCE(a.alumnos,0)+COALESCE(t.docentes,0)+COALESCE(v.visitas,0) total
        FROM users u
        LEFT JOIN (SELECT user_id, COUNT(*) alumnos FROM access_logs GROUP BY user_id) a ON a.user_id=u.id
        LEFT JOIN (SELECT user_id, COUNT(*) docentes FROM teacher_logs GROUP BY user_id) t ON t.user_id=u.id
        LEFT JOIN (SELECT user_id, COUNT(*) visitas FROM visitors GROUP BY user_id) v ON v.user_id=u.id
        ORDER BY total DESC, u.name ASC
    """).fetchall()
    recent = c.execute("""SELECT l.*, u.name user FROM access_logs l LEFT JOIN users u ON u.id=l.user_id ORDER BY l.id DESC LIMIT 8""").fetchall()
    recent_visitors = c.execute("""SELECT v.*, u.name user FROM visitors v LEFT JOIN users u ON u.id=v.user_id ORDER BY v.id DESC LIMIT 8""").fetchall()
    recent_teachers = c.execute("""SELECT t.*, u.name user FROM teacher_logs t LEFT JOIN users u ON u.id=t.user_id ORDER BY t.id DESC LIMIT 8""").fetchall()
    active_students_today = c.execute("""SELECT l.*, u.name user FROM access_logs l LEFT JOIN users u ON u.id=l.user_id
                                         WHERE l.result='ACTIVO' AND substr(l.created_at,1,10)=?
                                         ORDER BY l.id DESC LIMIT 12""", (today,)).fetchall()
    active_teachers_today = c.execute("""SELECT t.*, u.name user FROM teacher_logs t LEFT JOIN users u ON u.id=t.user_id
                                         WHERE t.result='ACTIVO' AND substr(t.created_at,1,10)=?
                                         ORDER BY t.id DESC LIMIT 12""", (today,)).fetchall()
    students = c.execute('SELECT * FROM students WHERE active=1 ORDER BY name LIMIT 8').fetchall()
    c.close()
    return render_template('dashboard.html', stats=stats, recent=recent, recent_visitors=recent_visitors,
                           recent_teachers=recent_teachers, students=students, weekly=weekly,
                           area_stats=area_stats, user_activity=user_activity,
                           sae_detail=sae_detail, sales_detail=sales_detail, security_detail=security_detail,
                           active_students_today=active_students_today, active_teachers_today=active_teachers_today,
                           lima_now=now_dt().strftime('%d/%m/%Y %H:%M'))

@app.route('/panel')
@login_required
@admin_required
def panel():
    return redirect(url_for('dashboard'))

@app.route('/alumnos')
@login_required
@role_required('admin','sae')
def students():
    q=request.args.get('q','').strip()
    c=conn()
    if q:
        data=c.execute("SELECT * FROM students WHERE name LIKE ? OR dni LIKE ? OR code LIKE ? ORDER BY active DESC, name",
            (f'%{q}%',f'%{q}%',f'%{q}%')).fetchall()
    else:
        data=c.execute('SELECT * FROM students ORDER BY active DESC, name').fetchall()
    c.close(); return render_template('students.html', students=data, q=q)


@app.route('/docentes')
@login_required
@role_required('admin','sae')
def teachers():
    q=request.args.get('q','').strip()
    c=conn()
    if q:
        data=c.execute("SELECT * FROM teachers WHERE name LIKE ? OR dni LIKE ? OR area LIKE ? ORDER BY active DESC, name",
            (f'%{q}%',f'%{q}%',f'%{q}%')).fetchall()
    else:
        data=c.execute('SELECT * FROM teachers ORDER BY active DESC, name').fetchall()
    c.close(); return render_template('teachers.html', teachers=data, q=q)


@app.route('/usuarios')
@login_required
@admin_required
def users():
    c=conn(); data=c.execute('SELECT id,name,email,role,created_at FROM users ORDER BY id DESC').fetchall(); c.close()
    return render_template('users.html', users=data)


@app.route('/registros')
@login_required
def logs():
    c=conn()
    if session.get('role') in ('admin','sae'):
        data=c.execute('''SELECT l.*, u.name user FROM access_logs l LEFT JOIN users u ON u.id=l.user_id ORDER BY l.id DESC''').fetchall()
    else:
        data=c.execute('''SELECT l.*, u.name user FROM access_logs l LEFT JOIN users u ON u.id=l.user_id WHERE l.user_id=? ORDER BY l.id DESC''',(session['user_id'],)).fetchall()
    c.close(); return render_template('logs.html', logs=data)


@app.route('/importar/alumnos', methods=['POST'])
@login_required
@admin_required
def import_excel():
    f=request.files.get('file')
    if not f or not f.filename.lower().endswith(('.xlsx','.xlsm')):
        flash('Sube un Excel válido .xlsx.', 'danger'); return redirect(url_for('dashboard'))
    path=os.path.join(UPLOADS, secure_filename(f.filename)); f.save(path)
    try:
        count=import_students_from_excel(path)
        flash(f'Base de alumnos actualizada: {count} activos. Los que no estén en el Excel quedan INACTIVOS.', 'success')
    except Exception as e:
        flash(str(e), 'danger')
    return redirect(url_for('dashboard'))


@app.route('/importar/docentes', methods=['POST'])
@login_required
@admin_required
def import_teachers():
    f=request.files.get('file')
    if not f or not f.filename.lower().endswith(('.xlsx','.xlsm')):
        flash('Sube un Excel válido .xlsx.', 'danger'); return redirect(url_for('teachers'))
    path=os.path.join(UPLOADS, secure_filename('docentes_' + f.filename)); f.save(path)
    try:
        count=import_teachers_from_excel(path)
        flash(f'Base de docentes actualizada: {count} activos. Los que no estén en el Excel quedan INACTIVOS.', 'success')
    except Exception as e:
        flash(str(e), 'danger')
    return redirect(url_for('teachers'))


def style_sheet(ws):
    fill = PatternFill('solid', fgColor='1D4ED8')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')
    for col in ws.columns:
        max_len = 12
        letter = col[0].column_letter
        for cell in col:
            max_len = max(max_len, len(str(cell.value or '')) + 2)
        ws.column_dimensions[letter].width = min(max_len, 35)


def build_report(kind='visitas', area=None):
    start = (now_dt() - timedelta(days=7)).isoformat(timespec='seconds')
    end = now_iso()
    c=conn()
    wb=Workbook()
    ws=wb.active
    if kind == 'visitas':
        ws.title='Visitas semanales'
        ws.append(['Fecha/Hora Lima', 'Registró', 'Nombres completos', 'DNI', 'Área destino', 'Motivo', 'Estado', 'Atendido en'])
        
        if area in AREAS:
            rows = c.execute('''SELECT v.*, u.name user FROM visitors v LEFT JOIN users u ON u.id=v.user_id
                                WHERE v.created_at BETWEEN ? AND ? AND v.destination_area=? ORDER BY v.created_at DESC''', (start,end,area)).fetchall()
        else:
            rows = c.execute('''SELECT v.*, u.name user FROM visitors v LEFT JOIN users u ON u.id=v.user_id
                                WHERE v.created_at BETWEEN ? AND ? ORDER BY v.created_at DESC''', (start,end)).fetchall()
        for r in rows:
            ws.append([r['created_at'], r['user'], r['full_name'], r['dni'], r['destination_area'], r['reason'], 'ATENDIDO' if r['attended'] else 'NO ATENDIDO', r['attended_at'] or ''])
        filename = f'reporte_visitas_semanal_lima_{now_dt().strftime("%Y%m%d_%H%M")}.xlsx'
    else:
        ws.title='Registros 7 dias'
        ws.append(['Fecha/Hora Lima','Usuario ID','Consulta','Resultado','Alumno','DNI','Codigo','Nota'])
        rows = c.execute('SELECT * FROM access_logs WHERE created_at BETWEEN ? AND ? ORDER BY id DESC', (start,end)).fetchall()
        for r in rows:
            ws.append([r['created_at'], r['user_id'], r['query'], r['result'], r['student_name'], r['student_dni'], r['student_code'], r['note']])
        filename = f'registros_ultimos_7_dias_lima_{now_dt().strftime("%Y%m%d_%H%M")}.xlsx'
    c.close()
    style_sheet(ws)
    path=os.path.join(UPLOADS, filename); wb.save(path)
    return path


@app.route('/exportar')
@login_required
@role_required('admin','sae')
def export_excel():
    return send_file(build_report('registros'), as_attachment=True)


@app.route('/exportar/visitas')
@login_required
@role_required('admin','sae','sales')
def export_visits_excel():
    area = None
    if session.get('role') == 'sae':
        area = 'SAE'
    elif session.get('role') == 'sales':
        area = 'VENTAS'
    return send_file(build_report('visitas', area=area), as_attachment=True)


if __name__ == '__main__':
    init_db()
    app.run(debug=True)
